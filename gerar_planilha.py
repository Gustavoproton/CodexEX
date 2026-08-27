#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador de Planilha de OS - divide os itens do escopo em Ativos, Consumíveis
e Gases (com acetileno sempre isolado), buscando código e valor no catálogo
mestre e preenchendo a tabela de itens no MODELO ORIGINAL da empresa.

O topo de cada aba (Solicitante, Cliente, CNPJ, OS, PO, Centro de Custo etc.)
fica em branco -- é preenchido manualmente depois, como no modelo enviado.

USO
----
    python gerar_planilha.py escopo.txt
    python gerar_planilha.py            # cola o escopo interativamente (Ctrl+D pra finalizar)
"""
import sys
import os
import re
import unicodedata

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from rapidfuzz import process, fuzz

CATALOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "catalogo_produtos.xlsx")
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_os.xlsx")
MATCH_THRESHOLD = 70


def normalizar(texto):
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return texto.upper().strip()


def carregar_catalogo(path=CATALOGO_PATH):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["PLANILHA_COMPLETA"]
    catalogo = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        categoria, codigo, descricao, und, valor = row[:5]
        if not descricao:
            continue
        catalogo.append({
            "categoria": (categoria or "").strip().upper(),
            "codigo": codigo,
            "descricao": str(descricao).strip(),
            "descricao_norm": normalizar(str(descricao)),
            "und": und or "UN",
            "valor": float(valor) if valor else 0.0,
        })
    return catalogo


NUM_RE = re.compile(r'\d+[.,]?\d*')


def extrair_numeros(texto):
    return set(NUM_RE.findall(texto))


def casar_item(desc_original, catalogo, descricoes_cache):
    consulta = normalizar(desc_original)
    numeros_consulta = extrair_numeros(consulta)
    candidatos = process.extract(consulta, descricoes_cache, scorer=fuzz.token_set_ratio, limit=6)

    qualificados = []
    for desc_cat, score, idx in candidatos:
        if score < MATCH_THRESHOLD:
            continue
        if numeros_consulta:
            numeros_cat = extrair_numeros(desc_cat)
            if not (numeros_consulta & numeros_cat):
                continue
        qualificados.append((idx, score))

    if qualificados:
        melhor_score = qualificados[0][1]
        empatados = [q for q in qualificados if abs(q[1] - melhor_score) < 0.01]
        if len(empatados) > 1:
            empatados = sorted(
                empatados,
                key=lambda q: fuzz.ratio(consulta, descricoes_cache[q[0]]),
                reverse=True,
            )
            melhor_secundario = fuzz.ratio(consulta, descricoes_cache[empatados[0][0]])
            reempatados = [
                q for q in empatados
                if abs(fuzz.ratio(consulta, descricoes_cache[q[0]]) - melhor_secundario) < 0.01
            ]
            idx, score = empatados[0]
            if len(reempatados) > 1:
                score = min(score, 90.0)
        else:
            idx, score = qualificados[0]
        return catalogo[idx], score

    if candidatos and not numeros_consulta and candidatos[0][1] >= 92:
        desc_cat, score, idx = candidatos[0]
        return catalogo[idx], score

    return None, candidatos[0][1] if candidatos else 0


LINE_RE = re.compile(r'''
    ^\s*
    (?:(?P<qtd_pre>\d+)\s*[xX]?\s*[-–]?\s*)?
    (?P<desc>.+?)
    (?:\s*[-–xX]\s*(?P<qtd_pos>\d+)\s*(?:un\.?|und\.?|unid\w*)?)?
    \s*$
''', re.VERBOSE)


def parse_escopo(texto):
    itens = []
    for linha in texto.splitlines():
        linha = linha.strip()
        if not linha:
            continue
        m = LINE_RE.match(linha)
        if not m:
            continue
        desc = (m.group("desc") or "").strip(" -–")
        qtd_str = m.group("qtd_pre") or m.group("qtd_pos") or "1"
        if desc:
            itens.append({"descricao_original": desc, "qtd": int(qtd_str)})
    return itens


def classificar(itens, catalogo):
    descricoes_cache = [c["descricao_norm"] for c in catalogo]
    grupos = {"ATIVO": [], "CONSUMIVEL": [], "GAS": []}
    nao_encontrados = []

    for item in itens:
        match, score = casar_item(item["descricao_original"], catalogo, descricoes_cache)
        if match is None:
            nao_encontrados.append({**item, "score": score})
            continue
        linha = {
            "codigo": match["codigo"],
            "descricao": match["descricao"],
            "und": match["und"],
            "qtd": item["qtd"],
            "valor_unit": match["valor"],
            "descricao_original": item["descricao_original"],
            "score": score,
        }
        categoria = match["categoria"] if match["categoria"] in grupos else "CONSUMIVEL"
        grupos[categoria].append(linha)

    gases = grupos.pop("GAS")
    acetileno = [g for g in gases if "ACETILENO" in g["descricao"].upper()]
    outros_gases = [g for g in gases if g not in acetileno]

    if acetileno and outros_gases:
        grupos["GAS"] = outros_gases
        grupos["ACETILENO"] = acetileno
    elif acetileno:
        grupos["ACETILENO"] = acetileno
        grupos["GAS"] = []
    else:
        grupos["GAS"] = outros_gases

    return grupos, nao_encontrados


PRIMEIRA_LINHA_ITEM = 19
COLUNAS_ITEM = ("A", "B", "G", "H", "I", "J")

CELULAS_CABECALHO_EM_BRANCO = [
    "B6", "H6", "A8", "F8", "B10", "B11", "G11",
    "C13", "H13", "C14", "H14", "C15", "H15", "C16", "J16",
]

FONT_TOTAL = Font(name="Arial", size=10, bold=True)
FILL_REVISAR = PatternFill("solid", fgColor="FFF2CC")


def _limpar_valor(ws, coord):
    ws[coord].value = None


def clonar_aba_modelo(wb, template_ws, nome_aba):
    ws = wb.copy_worksheet(template_ws)
    ws.title = nome_aba[:31]
    for coord in CELULAS_CABECALHO_EM_BRANCO:
        _limpar_valor(ws, coord)
    ultima_linha_modelo = template_ws.max_row
    for linha in range(PRIMEIRA_LINHA_ITEM, ultima_linha_modelo + 1):
        for col in COLUNAS_ITEM:
            _limpar_valor(ws, f"{col}{linha}")
    return ws, ultima_linha_modelo


def _copiar_estilo_linha(ws, linha_origem, linha_destino):
    for col in COLUNAS_ITEM:
        src = ws[f"{col}{linha_origem}"]
        dst = ws[f"{col}{linha_destino}"]
        dst.font = src.font.copy()
        dst.border = src.border.copy()
        dst.fill = src.fill.copy()
        dst.number_format = src.number_format
        dst.alignment = src.alignment.copy()
    ws.merge_cells(f"B{linha_destino}:F{linha_destino}")


def escrever_itens_no_modelo(ws, itens, ultima_linha_modelo):
    linha = PRIMEIRA_LINHA_ITEM
    for it in itens:
        if linha > ultima_linha_modelo:
            _copiar_estilo_linha(ws, ultima_linha_modelo, linha)

        revisar = it.get("score", 100) < 95
        fill = FILL_REVISAR if revisar else None

        ws[f"A{linha}"] = it["codigo"]
        desc = it["descricao"]
        if revisar:
            desc += f'  (conferir: digitado como "{it["descricao_original"]}")'
        ws[f"B{linha}"] = desc
        ws[f"G{linha}"] = it["und"]
        ws[f"H{linha}"] = it["qtd"]
        ws[f"I{linha}"] = it["valor_unit"]
        ws[f"I{linha}"].number_format = "#,##0.00"
        ws[f"J{linha}"] = f"=H{linha}*I{linha}"
        ws[f"J{linha}"].number_format = "#,##0.00"

        if fill:
            for col in COLUNAS_ITEM:
                ws[f"{col}{linha}"].fill = fill

        linha += 1

    linha_total = max(linha, ultima_linha_modelo + 1)
    if linha_total > ultima_linha_modelo:
        _copiar_estilo_linha(ws, ultima_linha_modelo, linha_total)
    ws[f"B{linha_total}"] = "TOTAL GERAL"
    ws[f"B{linha_total}"].font = FONT_TOTAL
    ws[f"B{linha_total}"].alignment = Alignment(horizontal="right")
    ws[f"J{linha_total}"] = f"=SUM(J{PRIMEIRA_LINHA_ITEM}:J{linha_total - 1})" if itens else 0
    ws[f"J{linha_total}"].font = FONT_TOTAL
    ws[f"J{linha_total}"].number_format = "#,##0.00"


def montar_aba(wb, template_ws, nome_aba, itens):
    ws, ultima_linha_modelo = clonar_aba_modelo(wb, template_ws, nome_aba)
    escrever_itens_no_modelo(ws, itens, ultima_linha_modelo)
    return ws


def montar_aba_nao_encontrados(wb, nao_encontrados):
    if not nao_encontrados:
        return
    ws = wb.create_sheet("Não encontrados")
    cab = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_cab = PatternFill("solid", fgColor="4472C4")
    ws["A1"] = "Descrição digitada no escopo"
    ws["B1"] = "Qtde"
    ws["C1"] = "Melhor score encontrado"
    for c in ("A1", "B1", "C1"):
        ws[c].font = cab
        ws[c].fill = fill_cab
    linha = 2
    for item in nao_encontrados:
        ws[f"A{linha}"] = item["descricao_original"]
        ws[f"B{linha}"] = item["qtd"]
        ws[f"C{linha}"] = round(item["score"], 1)
        linha += 1
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20


ARQUIVOS_SAIDA = [
    ("Ativos", [("ATIVO", "Ativos")]),
    ("Consumiveis", [("CONSUMIVEL", "Consumíveis"), ("GAS", "Gases")]),
    ("Acetileno", [("ACETILENO", "Acetileno")]),
]


def gerar_arquivo(template_path, sufixo_saida, categorias, grupos):
    grupos_com_item = [(chave, nome) for chave, nome in categorias if grupos.get(chave)]
    if not grupos_com_item:
        return None

    wb = openpyxl.load_workbook(template_path)
    template_ws = wb.active

    for chave, nome_aba in grupos_com_item:
        montar_aba(wb, template_ws, nome_aba, grupos[chave])

    wb.remove(template_ws)

    nome_arquivo = f"{sufixo_saida}.xlsx"
    wb.save(nome_arquivo)
    return nome_arquivo, [(nome, len(grupos[chave])) for chave, nome in grupos_com_item]


def main():
    if len(sys.argv) > 1:
        with open(sys.argv[1], "r", encoding="utf-8") as f:
            texto = f.read()
    else:
        print("Cole o escopo abaixo (um item por linha). Ctrl+D pra finalizar:")
        texto = sys.stdin.read()

    if not os.path.exists(TEMPLATE_PATH):
        print(f"ERRO: modelo não encontrado em {TEMPLATE_PATH}.")
        sys.exit(1)
    if not os.path.exists(CATALOGO_PATH):
        print(f"ERRO: catálogo não encontrado em {CATALOGO_PATH}.")
        sys.exit(1)

    catalogo = carregar_catalogo()
    itens = parse_escopo(texto)
    grupos, nao_encontrados = classificar(itens, catalogo)

    print()
    algum_arquivo = False
    for sufixo, categorias in ARQUIVOS_SAIDA:
        resultado = gerar_arquivo(TEMPLATE_PATH, sufixo, categorias, grupos)
        if resultado is None:
            if sufixo != "Acetileno":
                print(f"(nenhum item de {sufixo} encontrado - arquivo não gerado)")
            continue
        nome_arquivo, abas_info = resultado
        algum_arquivo = True
        print(f"Arquivo gerado: {nome_arquivo}")
        for nome_aba, qtd in abas_info:
            print(f"  - aba '{nome_aba}': {qtd} item(ns)")

    if nao_encontrados:
        wb_ne = openpyxl.Workbook()
        montar_aba_nao_encontrados(wb_ne, nao_encontrados)
        wb_ne.remove(wb_ne["Sheet"])
        wb_ne.save("Nao_encontrados.xlsx")
        print(f"Arquivo gerado: Nao_encontrados.xlsx - {len(nao_encontrados)} item(ns) pra revisar")

    if not algum_arquivo and not nao_encontrados:
        print("Nenhum item reconhecido no escopo. Confira o formato do arquivo de entrada.")

    print("\nLembrete: o topo de cada arquivo ficou em branco -> preencher manualmente.")


if __name__ == "__main__":
    main()
